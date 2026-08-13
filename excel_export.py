"""Excel-Export mit nachvollziehbaren Kosten-, Kurs- und Marge-Formeln."""

from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# SSI-nahe Farben (dezent, lesbar)
SSI_YELLOW = "FFED00"
SSI_BLACK = "1A1A1A"
SSI_GRAY = "F5F5F5"
SSI_HEADER = "2B2B2B"
SSI_MUTED = "666666"
SSI_GREEN = "E8F5E9"
SSI_BLUE = "E3F2FD"
SSI_ORANGE = "FFF3E0"

THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _safe_sheet_title(name: str) -> str:
    cleaned = re.sub(r"[\[\]\*\/\\\:\?]", "-", str(name or "Blatt"))[:31]
    return cleaned or "Blatt"


def _set_col_widths(ws: Worksheet, widths: Dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _style_title(cell) -> None:
    cell.font = Font(name="Calibri", size=16, bold=True, color=SSI_BLACK)
    cell.alignment = Alignment(vertical="center")


def _style_section(cell) -> None:
    cell.font = Font(name="Calibri", size=12, bold=True, color=SSI_BLACK)
    cell.fill = PatternFill("solid", fgColor=SSI_YELLOW)


def _style_header_row(ws: Worksheet, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=SSI_HEADER)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN
    ws.row_dimensions[row].height = 32


def _style_data_row(ws: Worksheet, row: int, start_col: int, end_col: int, alt: bool = False) -> None:
    fill = PatternFill("solid", fgColor=SSI_GRAY) if alt else None
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", size=10)
        cell.border = THIN
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill


def _money_format(cell, currency: str = "CHF") -> None:
    if currency.upper() == "EUR":
        cell.number_format = '#,##0.00 "EUR"'
    else:
        cell.number_format = '#,##0.00 "CHF"'


def _pct_format(cell) -> None:
    cell.number_format = '0.0"%"'


def _write_kv(ws: Worksheet, row: int, label: str, value: Any, value_fill: Optional[str] = None) -> int:
    ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=10, bold=True, color=SSI_MUTED)
    c = ws.cell(row=row, column=2, value=value)
    c.font = Font(name="Calibri", size=10)
    if value_fill:
        c.fill = PatternFill("solid", fgColor=value_fill)
    return row + 1


def _num(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _build_legend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Rechenweg"), 0)
    _set_col_widths(ws, {"A": 28, "B": 72, "C": 40})
    ws["A1"] = "Nachvollziehbarkeit – Formeln & Rechenweg"
    _style_title(ws["A1"])
    ws.merge_cells("A1:C1")

    rows = [
        [],
        ["Bereich", "Formel / Rechenweg", "Bemerkung"],
        [
            "Lizenzen – Einkauf CHF",
            "EK_CHF = IC_EUR × Kurs_EUR→CHF",
            "IC = Intercompany-Einkaufspreis in EUR",
        ],
        [
            "Lizenzen – Verkauf CHF",
            "VK_CHF = EK_CHF / (1 − Marge%/100), gerundet auf 10 CHF",
            "DB-Marge vom Verkaufspreis (wie Preisliste)",
        ],
        [
            "Lizenzen – Deckungsbeitrag",
            "DB_CHF = VK_CHF − EK_CHF",
            "DB% = DB_CHF / VK_CHF × 100",
        ],
        [
            "IT – Zeile Stunden",
            "Betrag = Stunden × Stundensatz (ggf. × (1+Zusatzmarge%))",
            "Standard Zusatz-Marge IT = 0%",
        ],
        [
            "IT – Material",
            "VK = EP / (1 − Materialmarge%/100)",
            "EP = Einkaufspreis Material",
        ],
        [
            "IT – interner DB (fiktiv)",
            "fiktiver EK = VK / 1.25  →  DB ≈ 20% vom VK",
            "Nur intern; nicht im Kundenangebot",
        ],
        [
            "Reisekosten",
            "Fahrzeit + km×Satz + Übernachtung + Verpflegung",
            "Basic 5/5 Fahrten/Verpflegung · Advanced 7/7",
        ],
        [],
        ["Farblegende", "", ""],
        ["Gelb", "Parameter / Eingaben (Kurs, Marge)", ""],
        ["Blau", "Einkauf / Kosten", ""],
        ["Grün", "Verkauf / Ergebnis", ""],
        ["Orange", "Zwischenschritte / Umrechnung", ""],
    ]
    for r in rows:
        ws.append(r)

    _style_header_row(ws, 3, 1, 3)
    for r in range(4, 11):
        _style_data_row(ws, r, 1, 3, alt=(r % 2 == 0))
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=SSI_ORANGE)

    fills = {
        13: SSI_YELLOW,
        14: SSI_BLUE,
        15: SSI_GREEN,
        16: SSI_ORANGE,
    }
    for r, color in fills.items():
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=color)
        _style_data_row(ws, r, 1, 3)

    ws.freeze_panes = "A4"


def _build_overview(wb: Workbook, offer: Dict[str, Any]) -> Worksheet:
    ws = wb.create_sheet(_safe_sheet_title("Uebersicht"), 1)
    _set_col_widths(ws, {"A": 32, "B": 28, "C": 18, "D": 18, "E": 40})
    meta = offer.get("meta") or {}
    customer = offer.get("customer") or {}
    kind = offer.get("kind") or "license"
    cfg = offer.get("configuration") or (offer.get("content") or {}).get("configurationSummary") or {}
    totals = offer.get("totals") or {}
    summary = offer.get("priceSummary") or {}

    ws["A1"] = "WAMAS Lift & Store – Kalkulationsübersicht"
    _style_title(ws["A1"])
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 28

    r = 3
    ws.cell(row=r, column=1, value="Stammdaten")
    _style_section(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    for label, value in [
        ("Angebotsart", kind),
        ("Nummer", meta.get("offerNumber")),
        ("Datum", meta.get("documentDate") or meta.get("createdAt")),
        ("Kunde", customer.get("company")),
        ("Projekt", customer.get("projectName")),
        ("Instanz", cfg.get("instanceName")),
        ("Geräte / Zonen / Öffnungen", f'{cfg.get("deviceCount") or "–"} / {cfg.get("zoneCount") or "–"} / {cfg.get("openingCount") or "–"}'),
    ]:
        r = _write_kv(ws, r, label, value)

    r += 1
    ws.cell(row=r, column=1, value="Kennzahlen (Ergebnis)")
    _style_section(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    if kind == "license":
        metrics = [
            ("IC Total EUR (Einkauf)", totals.get("net"), "EUR", SSI_BLUE),
            ("Kurs EUR→CHF", totals.get("eurToChfRate"), None, SSI_YELLOW),
            ("Einkauf CHF (IC×Kurs)", totals.get("costChf"), "CHF", SSI_BLUE),
            ("DB-Marge %", totals.get("marginPercent"), "pct", SSI_YELLOW),
            ("Verkauf CHF", totals.get("sellNetChf"), "CHF", SSI_GREEN),
            ("Deckungsbeitrag CHF", totals.get("contributionMarginChf"), "CHF", SSI_GREEN),
            ("DB % vom Verkauf", totals.get("contributionMarginPercent"), "pct", SSI_GREEN),
        ]
    elif kind == "it":
        metrics = [
            ("IT-Stunden", totals.get("workHours"), None, SSI_ORANGE),
            ("IT Einkauf/Basis CHF", totals.get("workAmountCost"), "CHF", SSI_BLUE),
            ("Reise Einkauf/Basis CHF", totals.get("travelAmountCost"), "CHF", SSI_BLUE),
            ("Zusatz-Marge %", totals.get("marginPercent"), "pct", SSI_YELLOW),
            ("Verkauf Total CHF", totals.get("totalAmount"), "CHF", SSI_GREEN),
            ("Fiktiver EK-Faktor", totals.get("internalCostFactor"), None, SSI_YELLOW),
            ("Fiktiver EK intern CHF", totals.get("impliedTotalCostChf"), "CHF", SSI_BLUE),
            ("DB intern CHF", totals.get("contributionMarginChf"), "CHF", SSI_GREEN),
            ("DB % vom Verkauf", totals.get("contributionMarginPercent"), "pct", SSI_GREEN),
        ]
    else:
        lic = summary.get("license") or {}
        it = summary.get("it") or {}
        mat = summary.get("material") or {}
        metrics = [
            ("Softwarelizenzen Verkauf CHF", lic.get("total"), "CHF", SSI_GREEN),
            ("IT-Aufwand Total CHF", it.get("total"), "CHF", SSI_GREEN),
            ("Material Total CHF", mat.get("total"), "CHF", SSI_GREEN),
            ("Gesamttotal CHF", summary.get("grandTotalChf"), "CHF", SSI_GREEN),
        ]

    for label, value, kind_fmt, fill in metrics:
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10, bold=True)
        cell = ws.cell(row=r, column=2, value=_num(value) if kind_fmt in {"CHF", "EUR", "pct", None} and value is not None and str(value).replace(".", "", 1).replace("-", "", 1).isdigit() or isinstance(value, (int, float)) else value)
        if isinstance(value, (int, float)):
            cell.value = float(value)
            if kind_fmt in {"CHF", "EUR"}:
                _money_format(cell, "EUR" if kind_fmt == "EUR" else "CHF")
            elif kind_fmt == "pct":
                _pct_format(cell)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.border = THIN
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Hinweis")
    _style_section(ws.cell(row=r, column=1))
    r += 1
    ws.cell(
        row=r,
        column=1,
        value=(
            "Die Blätter «Lizenzen» und «IT» enthalten Eingabewerte und Excel-Formeln. "
            "Endpreise sind nachvollziehbar über Kurs, Marge und Mengen. "
            "Interne DB-Faktoren sind gekennzeichnet und gehören nicht ins Kundenangebot."
        ),
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 48
    return ws


def _build_license_sheet(wb: Workbook, offer: Dict[str, Any]) -> None:
    lines = list(offer.get("lines") or [])
    optional = list(offer.get("optionalLines") or [])
    totals = offer.get("totals") or {}
    cfg = offer.get("configuration") or {}
    margin = _num(totals.get("marginPercent"), _num(cfg.get("licenseMarginPercent"), 28))
    rate = _num(totals.get("eurToChfRate"), _num(cfg.get("eurToChfRate"), 0.93))

    ws = wb.create_sheet(_safe_sheet_title("Lizenzen"))
    _set_col_widths(
        ws,
        {
            "A": 16, "B": 36, "C": 10, "D": 12, "E": 12, "F": 12,
            "G": 14, "H": 14, "I": 12, "J": 14, "K": 14, "L": 12, "M": 40,
        },
    )

    ws["A1"] = "Lizenzen – IC → Kurs → Einkauf → DB-Marge → Verkauf"
    _style_title(ws["A1"])
    ws.merge_cells("A1:M1")

    ws["A3"] = "Parameter"
    _style_section(ws["A3"])
    ws["A4"] = "Kurs EUR→CHF"
    ws["B4"] = rate
    ws["B4"].fill = PatternFill("solid", fgColor=SSI_YELLOW)
    ws["B4"].number_format = "0.0000"
    ws["C4"] = "← Zelle B4 (in Formeln: $B$4)"

    ws["A5"] = "DB-Marge %"
    ws["B5"] = margin
    ws["B5"].fill = PatternFill("solid", fgColor=SSI_YELLOW)
    _pct_format(ws["B5"])
    ws["C5"] = "← Zelle B5 (in Formeln: $B$5) · VK = EK/(1−m/100)"

    ws["A6"] = "Rundung VK"
    ws["B6"] = "auf 10 CHF (Excel: ROUND(…,-1))"
    ws["B6"].fill = PatternFill("solid", fgColor=SSI_ORANGE)

    headers = [
        "SKU",
        "Bezeichnung",
        "Menge",
        "IC EUR/Stk",
        "IC EUR Total",
        "Kurs",
        "EK CHF/Stk",
        "EK CHF Total",
        "Marge %",
        "VK CHF/Stk",
        "VK CHF Total",
        "DB CHF",
        "Beschreibung",
    ]
    header_row = 8
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, 1, len(headers))

    # Farbhinweis auf Header-Gruppen
    for col in range(4, 6):
        ws.cell(row=header_row, column=col).fill = PatternFill("solid", fgColor="1565C0")
    for col in range(6, 9):
        ws.cell(row=header_row, column=col).fill = PatternFill("solid", fgColor="0277BD")
    for col in range(9, 12):
        ws.cell(row=header_row, column=col).fill = PatternFill("solid", fgColor="2E7D32")
    ws.cell(row=header_row, column=12).fill = PatternFill("solid", fgColor="F9A825")

    def _append_lines(start_row: int, source: List[Dict[str, Any]], optional_flag: bool) -> int:
        r = start_row
        for idx, line in enumerate(source):
            qty = _num(line.get("qty"), 0)
            ic_unit = _num(line.get("unitPriceIcEur"), _num(line.get("unitPrice"), 0))
            # Falls unitPrice schon VK ist und unitPriceIcEur fehlt
            if line.get("unitPriceIcEur") is None and line.get("totalIcEur") is not None and qty:
                ic_unit = _num(line.get("totalIcEur")) / qty if qty else _num(line.get("totalIcEur"))

            ws.cell(row=r, column=1, value=line.get("sku") or "")
            name = str(line.get("name") or "")
            if optional_flag:
                name = f"[Option] {name}"
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value=qty)

            c_ic = ws.cell(row=r, column=4, value=ic_unit)
            _money_format(c_ic, "EUR")
            c_ic.fill = PatternFill("solid", fgColor=SSI_BLUE)

            # IC Total = Menge × IC/Stk
            c_ic_tot = ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
            _money_format(c_ic_tot, "EUR")
            c_ic_tot.fill = PatternFill("solid", fgColor=SSI_BLUE)

            c_rate = ws.cell(row=r, column=6, value="=$B$4")
            c_rate.number_format = "0.0000"
            c_rate.fill = PatternFill("solid", fgColor=SSI_YELLOW)

            # EK/Stk = IC × Kurs
            c_ek = ws.cell(row=r, column=7, value=f"=D{r}*F{r}")
            _money_format(c_ek, "CHF")
            c_ek.fill = PatternFill("solid", fgColor=SSI_BLUE)

            c_ek_tot = ws.cell(row=r, column=8, value=f"=C{r}*G{r}")
            _money_format(c_ek_tot, "CHF")
            c_ek_tot.fill = PatternFill("solid", fgColor=SSI_BLUE)

            c_m = ws.cell(row=r, column=9, value="=$B$5")
            _pct_format(c_m)
            c_m.fill = PatternFill("solid", fgColor=SSI_YELLOW)

            # VK/Stk = ROUND(EK/(1-m/100), -1)
            c_vk = ws.cell(
                row=r,
                column=10,
                value=f'=IF($B$5>=100,ROUND(G{r},-1),ROUND(G{r}/(1-$B$5/100),-1))',
            )
            _money_format(c_vk, "CHF")
            c_vk.fill = PatternFill("solid", fgColor=SSI_GREEN)

            c_vk_tot = ws.cell(row=r, column=11, value=f"=C{r}*J{r}")
            _money_format(c_vk_tot, "CHF")
            c_vk_tot.fill = PatternFill("solid", fgColor=SSI_GREEN)

            c_db = ws.cell(row=r, column=12, value=f"=K{r}-H{r}")
            _money_format(c_db, "CHF")
            c_db.fill = PatternFill("solid", fgColor=SSI_ORANGE)

            ws.cell(row=r, column=13, value=line.get("description") or "")
            _style_data_row(ws, r, 1, 13, alt=(idx % 2 == 1))
            r += 1
        return r

    row = header_row + 1
    first_data = row
    row = _append_lines(row, lines, False)
    if optional:
        ws.cell(row=row, column=1, value="Optionale Positionen (nicht im Total)")
        _style_section(ws.cell(row=row, column=1))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        row = _append_lines(row, optional, True)

    last_data = row - 1
    firm_end = header_row + len(lines) if lines else header_row
    if lines and firm_end >= first_data:
        row += 1
        ws.cell(row=row, column=1, value="Summen (Festpositionen)")
        _style_section(ws.cell(row=row, column=1))
        row += 1
        ws.cell(row=row, column=7, value="Σ EK CHF")
        cell_ek = ws.cell(row=row, column=8, value=f"=SUM(H{first_data}:H{firm_end})")
        _money_format(cell_ek, "CHF")
        cell_ek.fill = PatternFill("solid", fgColor=SSI_BLUE)
        sum_ek_row = row
        row += 1
        ws.cell(row=row, column=7, value="Σ VK CHF")
        cell_vk = ws.cell(row=row, column=11, value=f"=SUM(K{first_data}:K{firm_end})")
        _money_format(cell_vk, "CHF")
        cell_vk.fill = PatternFill("solid", fgColor=SSI_GREEN)
        sum_vk_row = row
        row += 1
        ws.cell(row=row, column=7, value="Σ DB CHF")
        cell_db = ws.cell(row=row, column=12, value=f"=K{sum_vk_row}-H{sum_ek_row}")
        _money_format(cell_db, "CHF")
        cell_db.fill = PatternFill("solid", fgColor=SSI_ORANGE)
        sum_db_row = row
        row += 1
        ws.cell(row=row, column=7, value="DB %")
        cell_pct = ws.cell(
            row=row,
            column=12,
            value=f'=IF(K{sum_vk_row}=0,0,L{sum_db_row}/K{sum_vk_row}*100)',
        )
        _pct_format(cell_pct)
        cell_pct.fill = PatternFill("solid", fgColor=SSI_GREEN)

        if totals.get("discountSellChf") or totals.get("discountAmount"):
            row += 2
            ws.cell(row=row, column=1, value="SLL-Rabatt (falls aktiv)")
            _style_section(ws.cell(row=row, column=1))
            row += 1
            r = _write_kv(ws, row, "IC-Rabatt EUR", totals.get("discountAmount"), SSI_BLUE)
            r = _write_kv(ws, r, "Verkaufs-Rabatt CHF", totals.get("discountSellChf"), SSI_GREEN)
            _write_kv(ws, r, "Verkauf netto CHF", totals.get("sellNetChf"), SSI_GREEN)

    ws.freeze_panes = "A9"
    if last_data >= header_row:
        ws.auto_filter.ref = f"A{header_row}:M{max(header_row, last_data)}"


def _build_it_sheet(wb: Workbook, offer: Dict[str, Any]) -> None:
    lines = list(offer.get("lines") or [])
    totals = offer.get("totals") or {}
    cfg = offer.get("configuration") or {}
    hourly = _num(cfg.get("hourlyRate"), 210)
    margin = _num(totals.get("marginPercent"), _num(cfg.get("itMarginPercent"), 0))
    factor = _num(totals.get("internalCostFactor"), 1.25)

    ws = wb.create_sheet(_safe_sheet_title("IT"))
    _set_col_widths(
        ws,
        {
            "A": 14, "B": 28, "C": 36, "D": 12, "E": 10, "F": 12,
            "G": 12, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14,
        },
    )

    ws["A1"] = "IT-Aufwand – Basis/Einkauf → Zusatzmarge → Verkauf → interner DB"
    _style_title(ws["A1"])
    ws.merge_cells("A1:L1")

    ws["A3"] = "Parameter"
    _style_section(ws["A3"])
    ws["A4"] = "Stundensatz CHF"
    ws["B4"] = hourly
    _money_format(ws["B4"], "CHF")
    ws["B4"].fill = PatternFill("solid", fgColor=SSI_YELLOW)

    ws["A5"] = "Zusatz-Marge % (auf IT)"
    ws["B5"] = margin
    _pct_format(ws["B5"])
    ws["B5"].fill = PatternFill("solid", fgColor=SSI_YELLOW)

    ws["A6"] = "Interner EK-Faktor (VK÷x)"
    ws["B6"] = factor
    ws["B6"].number_format = "0.00"
    ws["B6"].fill = PatternFill("solid", fgColor=SSI_YELLOW)
    ws["C6"] = "nur intern · fiktiver EK = VK / Faktor"

    headers = [
        "SKU",
        "Bezeichnung",
        "Beschreibung",
        "Kategorie",
        "Menge",
        "Stunden",
        "Basis CHF",
        "Zusatzmarge %",
        "Verkauf CHF",
        "DB Zeile CHF",
        "fikt. EK intern",
        "DB intern CHF",
    ]
    header_row = 8
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, 1, len(headers))

    row = header_row + 1
    first = row
    for idx, line in enumerate(lines):
        cat = str(line.get("category") or "")
        is_material = cat == "material" and line.get("billing") == "material"
        basis = _num(line.get("amountCost"), _num(line.get("amount"), 0))
        # For pre-margin snapshot amountCost is basis; amount is sell
        if line.get("amountCost") is None:
            # reconstruct: if margin applied, amount is sell
            basis = _num(line.get("amount"), 0)
            if margin and not is_material:
                basis = round(basis / (1 + margin / 100.0), 2)

        ws.cell(row=row, column=1, value=line.get("sku") or "")
        ws.cell(row=row, column=2, value=line.get("name") or "")
        ws.cell(row=row, column=3, value=line.get("description") or "")
        ws.cell(row=row, column=4, value=cat)
        ws.cell(row=row, column=5, value=_num(line.get("qty"), 0))
        ws.cell(row=row, column=6, value=_num(line.get("hours"), 0))

        c_basis = ws.cell(row=row, column=7, value=basis)
        _money_format(c_basis, "CHF")
        c_basis.fill = PatternFill("solid", fgColor=SSI_BLUE)

        if is_material:
            # Material: Basis = EP, Verkauf bereits mit Materialmarge – show stored sell, margin from values
            sell = _num(line.get("amount"), basis)
            c_m = ws.cell(row=row, column=8, value="")
            c_m.fill = PatternFill("solid", fgColor=SSI_ORANGE)
            c_sell = ws.cell(row=row, column=9, value=sell)
            _money_format(c_sell, "CHF")
            c_sell.fill = PatternFill("solid", fgColor=SSI_GREEN)
            c_db = ws.cell(row=row, column=10, value=f"=I{row}-G{row}")
            _money_format(c_db, "CHF")
            # Material: fiktiver EK = echter EP
            c_fek = ws.cell(row=row, column=11, value=f"=G{row}")
            _money_format(c_fek, "CHF")
            c_fek.fill = PatternFill("solid", fgColor=SSI_BLUE)
        else:
            c_m = ws.cell(row=row, column=8, value="=$B$5")
            _pct_format(c_m)
            c_m.fill = PatternFill("solid", fgColor=SSI_YELLOW)
            c_sell = ws.cell(row=row, column=9, value=f"=G{row}*(1+$B$5/100)")
            _money_format(c_sell, "CHF")
            c_sell.fill = PatternFill("solid", fgColor=SSI_GREEN)
            c_db = ws.cell(row=row, column=10, value=f"=I{row}-G{row}")
            _money_format(c_db, "CHF")
            c_fek = ws.cell(row=row, column=11, value=f"=IF($B$6=0,0,I{row}/$B$6)")
            _money_format(c_fek, "CHF")
            c_fek.fill = PatternFill("solid", fgColor=SSI_BLUE)

        c_idb = ws.cell(row=row, column=12, value=f"=I{row}-K{row}")
        _money_format(c_idb, "CHF")
        c_idb.fill = PatternFill("solid", fgColor=SSI_ORANGE)

        _style_data_row(ws, row, 1, 12, alt=(idx % 2 == 1))
        row += 1

    last = row - 1
    if last >= first:
        row += 1
        ws.cell(row=row, column=1, value="Summen")
        _style_section(ws.cell(row=row, column=1))
        row += 1
        labels = [
            (7, f"=SUM(G{first}:G{last})", "Σ Basis CHF", SSI_BLUE),
            (9, f"=SUM(I{first}:I{last})", "Σ Verkauf CHF", SSI_GREEN),
            (10, f"=SUM(J{first}:J{last})", "Σ DB Zeile", SSI_ORANGE),
            (11, f"=SUM(K{first}:K{last})", "Σ fikt. EK", SSI_BLUE),
            (12, f"=SUM(L{first}:L{last})", "Σ DB intern", SSI_GREEN),
        ]
        ws.cell(row=row, column=6, value="Summen →")
        for col, formula, label, color in labels:
            ws.cell(row=row, column=col - 1 if col > 7 else 5, value=label)
            cell = ws.cell(row=row, column=col, value=formula)
            _money_format(cell, "CHF")
            cell.fill = PatternFill("solid", fgColor=color)
        # cleaner summary block
        row += 2
        ws.cell(row=row, column=1, value="Kontrollwerte aus Angebot")
        _style_section(ws.cell(row=row, column=1))
        row += 1
        for label, val, fill in [
            ("Verkauf Total (JSON)", totals.get("totalAmount"), SSI_GREEN),
            ("Fiktiver EK intern (JSON)", totals.get("impliedTotalCostChf"), SSI_BLUE),
            ("DB intern (JSON)", totals.get("contributionMarginChf"), SSI_GREEN),
            ("DB % (JSON)", totals.get("contributionMarginPercent"), SSI_GREEN),
        ]:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=_num(val) if val is not None else None)
            if isinstance(val, (int, float)):
                if " %" in label or label.startswith("DB %"):
                    _pct_format(cell)
                else:
                    _money_format(cell, "CHF")
            cell.fill = PatternFill("solid", fgColor=fill)
            row += 1

    ws.freeze_panes = "A9"
    if last >= first:
        ws.auto_filter.ref = f"A{header_row}:L{last}"


def _build_offer_document_sheet(wb: Workbook, offer: Dict[str, Any]) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Gesamtangebot"))
    _set_col_widths(
        ws,
        {"A": 8, "B": 22, "C": 16, "D": 28, "E": 40, "F": 10, "G": 14, "H": 10, "I": 14, "J": 10},
    )
    summary = offer.get("priceSummary") or {}
    ws["A1"] = "Gesamtangebot – Positionen (Verkaufspreise)"
    _style_title(ws["A1"])
    ws.merge_cells("A1:J1")
    ws["A2"] = (
        "Hinweis: Detail-Rechenwege zu Lizenzen/IT stehen in den Blättern «Lizenzen» und «IT», "
        "sofern die zugehörigen Quellenkalkulationen im Angebot verknüpft sind. "
        "Hier: kundenrelevante Verkaufspositionen."
    )
    ws.merge_cells("A2:J2")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 36

    headers = ["Pos", "Bereich", "SKU", "Bezeichnung", "Beschreibung", "Menge", "Einzelpreis", "Stunden", "Betrag", "Währung"]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, 1, len(headers))

    row = header_row + 1
    first = row
    for idx, line in enumerate(offer.get("commercialLines") or []):
        vals = [
            line.get("pos"),
            line.get("section"),
            line.get("sku"),
            line.get("name"),
            line.get("description"),
            _num(line.get("qty"), 0) if line.get("qty") is not None else None,
            _num(line.get("unitPrice")) if line.get("unitPrice") is not None else None,
            _num(line.get("hours")) if line.get("hours") is not None else None,
            _num(line.get("amount")) if line.get("amount") is not None else None,
            line.get("currency") or "CHF",
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            if c in {7, 9} and isinstance(v, (int, float)):
                _money_format(cell, str(vals[9] or "CHF"))
        _style_data_row(ws, row, 1, 10, alt=(idx % 2 == 1))
        ws.cell(row=row, column=9).fill = PatternFill("solid", fgColor=SSI_GREEN)
        row += 1
    last = row - 1

    row += 1
    ws.cell(row=row, column=1, value="Zusammenfassung")
    _style_section(ws.cell(row=row, column=1))
    row += 1
    lic = summary.get("license") or {}
    it = summary.get("it") or {}
    mat = summary.get("material") or {}
    for label, val in [
        ("Softwarelizenzen Verkauf CHF", lic.get("total")),
        ("IT-Aufwand Total CHF", it.get("total")),
        ("Material Total CHF", mat.get("total")),
        ("Gesamttotal CHF", summary.get("grandTotalChf")),
    ]:
        ws.cell(row=row, column=8, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=9, value=_num(val) if val is not None else None)
        _money_format(cell, "CHF")
        cell.fill = PatternFill("solid", fgColor=SSI_GREEN)
        row += 1
    if summary.get("note"):
        row += 1
        ws.cell(row=row, column=1, value=str(summary.get("note")))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    if last >= first:
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A{header_row}:J{last}"


def build_offer_excel(offer: Dict[str, Any]) -> bytes:
    """Erzeugt formatierte XLSX mit Rechenweg, Parametern und Excel-Formeln."""
    wb = Workbook()
    # remove default sheet; we create ordered sheets
    default = wb.active
    wb.remove(default)

    kind = offer.get("kind") or "license"
    _build_legend_sheet(wb)
    _build_overview(wb, offer)

    if kind == "license":
        _build_license_sheet(wb, offer)
    elif kind == "it":
        _build_it_sheet(wb, offer)
    elif kind == "offer_document":
        _build_offer_document_sheet(wb, offer)
        # Wenn eingebettete Quellen existieren, Detailblätter ergänzen
        editable = offer.get("editable") or {}
        lic = editable.get("license")
        it = editable.get("it")
        if isinstance(lic, dict) and lic.get("lines"):
            _build_license_sheet(wb, lic)
        if isinstance(it, dict) and it.get("lines"):
            _build_it_sheet(wb, it)
    else:
        # Fallback: Rohdaten
        ws = wb.create_sheet("Daten")
        ws.append(["Unbekannte Angebotsart", kind])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()
